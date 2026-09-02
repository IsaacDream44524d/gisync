import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from app.models.students import Group, Student

def generate_groups_excel_buffer(db_session) -> io.BytesIO:
    """Generates the Excel file in-memory and returns a BytesIO buffer."""
    wb = Workbook()

    # Sheet 1: Groups View
    ws_groups = wb.active
    ws_groups.title = "Groups"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    group_title_font = Font(bold=True, size=12)

    groups = db_session.scalars(select(Group).order_by(Group.name)).all()

    for group in groups:
        ws_groups.append([group.name, f"{len(group.students)} Students"])
        ws_groups.cell(row=ws_groups.max_row, column=1).font = group_title_font

        ws_groups.append(["Name", "Email", "Gender", "Year"])
        header_row = ws_groups.max_row
        for col in range(1, 5):
            cell = ws_groups.cell(row=header_row, column=col)
            cell.font = header_font
            cell.fill = header_fill

        if group.students:
            for student in group.students:
                gender_val = student.gender.value if hasattr(student.gender, 'value') else (student.gender or "")
                ws_groups.append([
                    getattr(student, 'name', getattr(student, 'username', '')),
                    student.email,
                    gender_val,
                    getattr(student, 'year', '')
                ])
        else:
            ws_groups.append(["No students assigned", "", "", ""])

        ws_groups.append([])
        ws_groups.append([])

    # Sheet 2: Master Student View
    ws_students = wb.create_sheet("Students")
    ws_students.append(["Name", "Email", "Gender", "Year", "Group"])
    
    for col in range(1, 6):
        cell = ws_students.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    students = db_session.scalars(select(Student)).all()

    for student in students:
        gender_val = student.gender.value if hasattr(student.gender, 'value') else (student.gender or "")
        ws_students.append([
            getattr(student, 'name', getattr(student, 'username', '')),
            student.email,
            gender_val,
            getattr(student, 'year', ''),
            student.group.name if student.group else "Unassigned"
        ])

    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer