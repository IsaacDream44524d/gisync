import openpyxl
from email_validator import validate_email, EmailNotValidError


REQUIRED_HEADERS = ['fullname', 'email', 'year']

def extract_students(file):
    #load file
    try:
        workbook = openpyxl.load_workbook(file, read_only=True)
        #get the required sheet
    except FileNotFoundError:
        return 'file not found'

    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(cell).strip().lower() if cell else '' for cell in header_row]

    if headers != REQUIRED_HEADERS:
        raise ValueError(
            f"Invalid spreadsheet headers.\n"
            f"Expected: {REQUIRED_HEADERS}\n"
            f"Found: {headers}"
        )

    students = []
    errors = []
    seen_emails = set()

    for row_number, row  in enumerate(sheet.iter_rows(min_row=2, max_col=4, values_only=True)):
        if not row  or all(cell is None for cell in row):
            continue

        #headers
        fullname = str(row[0].strip().lower() if row[0] else '')
        email = str(row[1].strip().lower() if row[1] else '')
        year = row[2]

        if not fullname:
            errors.append(
                f"Row {row_number}: 'fullname' is required."
            )
            continue

        if not email:
            errors.append(
                f"Row {row_number}: 'email' is required."
            )
            continue

        if not year:
            errors.append(
                f"Row {row_number}: 'year' is required."
            )
            continue

        #email validation
        try:
            validate_email(email, check_deliverability=False)

        except EmailNotValidError:
            errors.append(
                f"Row {row_number}: Invalid email '{email}. Ignored"
            )
            continue

        if email in seen_emails:
            errors.append(
                f"Row {row_number}: Duplicate email {email}., ignored the last one"
            )
            continue

        seen_emails.add(email)

        #year validation
        try:
            year = int(year)

        except (ValueError, TypeError):
            errors.append(
                f"Row {row_number}: Invalid year"
            )

        if year not in (1, 2, 3, 4, 5):
            errors.append(
                f"Row {row_number}: Year must be between 1 and 5. '{year}'"
            )
            continue

        students.append({
            'fullname': fullname,
            'email': email,
            'year': year
        })

    workbook.close()
    return students, errors

def extract_timetable(file):
    pass
