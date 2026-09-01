from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select

def export_groups_to_excel(session, filepath: str):
    wb = Workbook()

    # ==========================
    # GROUPS SHEET
    # ==========================
    ws_groups = wb.active
    ws_groups.title = "Groups"

    groups = session.scalars(
        select(Group).order_by(Group.name)
    ).all()

    row = 1

    for group in groups:

        ws_groups.cell(row=row, column=1, value=group.name)
        ws_groups.cell(row=row, column=1).font = Font(bold=True)

        ws_groups.cell(
            row=row,
            column=2,
            value=f"{len(group.students)} Students"
        )

        row += 1

        ws_groups.append([
            "Name",
            "Email",
            "Gender",
            "Year"
        ])

        row += 1

        for student in group.students:
            ws_groups.append([
                student.username,
                student.email,
                student.gender.value if student.gender else "",
                student.year
            ])
            row += 1

        row += 2

    # ==========================
    # STUDENTS SHEET
    # ==========================
    ws_students = wb.create_sheet("Students")

    ws_students.append([
        "Name",
        "Email",
        "Gender",
        "Year",
        "Group"
    ])

    students = session.scalars(
        select(Student)
        .order_by(Student.username)
    ).all()

    for student in students:
        ws_students.append([
            student.username,
            student.email,
            student.gender.value if student.gender else "",
            student.year,
            student.group.name if student.group else ""
        ])

    # Auto-size columns
    for sheet in wb.worksheets:
        for column in sheet.columns:
            length = max(
                len(str(cell.value or ""))
                for cell in column
            )

            sheet.column_dimensions[
                column[0].column_letter
            ].width = min(length + 4, 40)

    wb.save(filepath)